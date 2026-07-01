"""Import Employee_Database and Attendance from the HR Excel file."""
import json
import re
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "HR System June 2026 V.2 (1).xlsx"
OUT = ROOT / "hr-system" / "data"

STATUS_COL_START = 16  # column P


def parse_attendance_dates(ws):
    """Read date headers from row 2 (Status columns)."""
    dates = []
    for col in range(STATUS_COL_START, ws.max_column + 1, 2):
        header = ws.cell(1, col).value
        if header != "Status":
            continue
        val = ws.cell(2, col).value
        if isinstance(val, datetime):
            dates.append(val.date().isoformat())
        elif val:
            try:
                dates.append(str(val)[:10])
            except Exception:
                pass
    return dates


def normalize_status(raw):
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    mapping = {
        "lateness a": "Lateness A",
        "lateness b": "Lateness B",
        "lateness A": "Lateness A",
        "lateness B": "Lateness B",
    }
    return mapping.get(s, s)


def import_employees(wb):
    ws = wb["Employee_Database"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    employees = []
    for row in range(2, ws.max_row + 1):
        emp_id = ws.cell(row, 1).value
        if not emp_id:
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            key = re.sub(r"[^a-zA-Z0-9]+", "_", str(h)).strip("_").lower()
            val = ws.cell(row, i + 1).value
            if isinstance(val, datetime):
                val = val.date().isoformat()
            rec[key] = val
        rec["id"] = str(emp_id).strip()
        employees.append(rec)
    return employees


def import_attendance_month(wb, sheet_name="Attendance"):
    ws = wb[sheet_name]
    dates = parse_attendance_dates(ws)
    if not dates:
        return [], []

    # infer year-month from first date
    first = dates[0]
    year_month = first[:7]

    records = []
    for row in range(3, ws.max_row + 1):
        emp_id = ws.cell(row, 1).value
        if not emp_id:
            continue
        emp_id = str(emp_id).strip()
        unit = ws.cell(row, 2).value
        name = ws.cell(row, 3).value
        email = ws.cell(row, 4).value

        col_idx = 0
        for col in range(STATUS_COL_START, ws.max_column + 1, 2):
            if col_idx >= len(dates):
                break
            d = dates[col_idx]
            status = normalize_status(ws.cell(row, col).value)
            fp = ws.cell(row, col + 1).value
            if status or fp:
                records.append({
                    "employeeId": emp_id,
                    "date": d,
                    "status": status or "Attended",
                    "fpLateness": str(fp).strip() if fp else None,
                    "unit": unit,
                    "name": name,
                    "email": email,
                })
            col_idx += 1

    return year_month, records


def month_working_days(year: int, month: int) -> int:
    """Count Mon-Fri days in month (default working calendar)."""
    d = date(year, month, 1)
    count = 0
    while d.month == month:
        if d.weekday() < 5:
            count += 1
        d += timedelta(days=1)
    return count


def build_july_from_employees(employees, june_records, active_statuses=None):
    """Initialize July 2026 attendance from employee list + June employee IDs."""
    if active_statuses is None:
        active_statuses = {
            "Active", "OUT BUT STILL GET PAID", "Paused", "Paused still get paid"
        }

    june_ids = {r["employeeId"] for r in june_records}
    year, month = 2026, 7
    records = []
    seen = set()

    def add_emp(emp):
        eid = emp["id"]
        if eid in seen:
            return
        seen.add(eid)
        d = date(year, month, 1)
        while d.month == month:
            ds = d.isoformat()
            dow = d.weekday()  # 5=Sat, 6=Sun
            if dow >= 5:
                records.append({
                    "employeeId": eid,
                    "date": ds,
                    "status": "Day-OFF",
                    "fpLateness": None,
                    "isWeekendDefault": True,
                    "unit": emp.get("unit"),
                    "name": emp.get("american_name") or emp.get("arabic_name"),
                    "email": emp.get("email"),
                })
            d += timedelta(days=1)

    for emp in employees:
        status = emp.get("status")
        eid = emp["id"]
        if status in active_statuses or eid in june_ids:
            add_emp(emp)

    return f"{year}-{month:02d}", records


def import_position_rates(wb):
    ws = wb["Position_Rates"]
    rates = []
    for row in range(4, ws.max_row + 1):
        pos = ws.cell(row, 1).value
        salary = ws.cell(row, 2).value
        if pos and salary is not None:
            rates.append({"position": str(pos).strip(), "monthlySalary": float(salary)})
    return rates


def import_bonus_deduction_sheet(wb, sheet_name, default_type):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    records = []
    seen = set()
    for row in range(2, ws.max_row + 1):
        emp_id = ws.cell(row, 1).value
        amount = ws.cell(row, 5).value
        if not emp_id or amount is None:
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            continue
        if amt == 0:
            continue
        date_val = ws.cell(row, 4).value
        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        elif date_val:
            date_str = str(date_val)[:10]
        else:
            date_str = "2026-06-01"
        bonus_type = str(ws.cell(row, 7).value or default_type).strip()
        key = (str(emp_id).strip(), date_str, bonus_type, amt)
        if key in seen:
            continue
        seen.add(key)
        records.append({
            "employeeId": str(emp_id).strip(),
            "date": date_str,
            "amount": amt,
            "reason": str(ws.cell(row, 6).value or "").strip(),
            "type": bonus_type,
            "unit": str(ws.cell(row, 2).value or "").strip(),
        })
    return records


def import_all_bonuses(wb):
    records = []
    seen = set()
    for name in wb.sheetnames:
        if name == "Bonus" or name.startswith("Bonus_"):
            for rec in import_bonus_deduction_sheet(wb, name, "Other Bonus"):
                key = (rec["employeeId"], rec["date"], rec["type"], rec["amount"])
                if key not in seen:
                    seen.add(key)
                    records.append(rec)
    return records


def import_all_deductions(wb):
    records = []
    seen = set()
    for name in wb.sheetnames:
        if name == "Deduction" or name.startswith("Deduction_"):
            for rec in import_bonus_deduction_sheet(wb, name, "Other Deductions"):
                key = (rec["employeeId"], rec["date"], rec["type"], rec["amount"])
                if key not in seen:
                    seen.add(key)
                    records.append(rec)
    return records


def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, str) and val.strip().startswith("#"):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def import_commission_types(wb):
    if "Commission_Types" not in wb.sheetnames:
        return []
    ws = wb["Commission_Types"]
    types = []
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 1).value
        rate = ws.cell(row, 2).value
        if not name:
            continue
        types.append({
            "name": str(name).strip(),
            "rateEgp": safe_float(rate),
            "description": str(ws.cell(row, 3).value or "").strip(),
            "active": str(ws.cell(row, 4).value or "Yes").lower() in ("yes", "true", "1"),
        })
    return types


def import_payroll_adjustments(wb, sheet_name="Payroll_June2026", year_month="2026-06"):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    records = []
    for row in range(5, ws.max_row + 1):
        emp_id = ws.cell(row, 1).value
        if not emp_id:
            continue
        extra = safe_float(ws.cell(row, 14).value)
        hold_raw = ws.cell(row, 59).value
        two_week_hold = str(hold_raw or "").strip().lower() in ("yes", "true", "1")
        commission_type = ws.cell(row, 56).value
        commission_amount = safe_float(ws.cell(row, 57).value)
        commission_comments = str(ws.cell(row, 58).value or "").strip()
        if not (extra or two_week_hold or commission_type or commission_amount):
            continue
        records.append({
            "employeeId": str(emp_id).strip(),
            "yearMonth": year_month,
            "extraDays": extra,
            "twoWeekHold": two_week_hold,
            "commissionType": str(commission_type).strip() if commission_type else "",
            "commissionAmount": commission_amount,
            "commissionComments": commission_comments,
        })
    return records


def import_bonuses(wb):
    return import_all_bonuses(wb)


def import_deductions(wb):
    return import_all_deductions(wb)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    employees = import_employees(wb)
    june_month, june_records = import_attendance_month(wb)
    july_month, july_records = build_july_from_employees(employees, june_records)
    rates = import_position_rates(wb)
    bonuses = import_bonuses(wb)
    deductions = import_deductions(wb)
    commission_types = import_commission_types(wb)
    payroll_adjustments = import_payroll_adjustments(wb, "Payroll_June2026", june_month)

    # Merge June imported records into store (historical)
    all_attendance = {}
    if june_month and june_records:
        all_attendance[june_month] = june_records
    all_attendance[july_month] = july_records

    config = {
        "defaultWeekendDays": [6, 0],  # Sunday=6 in JS, Saturday=6... Python: Sat=5 Sun=6
        "weekendDayNames": ["Saturday", "Sunday"],
        "latenessRules": {
            "tierA": {"label": "Lateness A", "beforeHour": 15, "amount": 25},
            "tierB": {"label": "Lateness B", "afterHour": 15, "amount": 50},
        },
        "workingDaysByMonth": {
            june_month: 22,
            july_month: month_working_days(2026, 7),
        },
    }

    (OUT / "employees.json").write_text(
        json.dumps(employees, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "attendance.json").write_text(
        json.dumps(all_attendance, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "position_rates.json").write_text(
        json.dumps(rates, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "config.json").write_text(
        json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "bonuses.json").write_text(
        json.dumps(bonuses, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "deductions.json").write_text(
        json.dumps(deductions, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "commission_types.json").write_text(
        json.dumps(commission_types, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "payroll_adjustments.json").write_text(
        json.dumps(payroll_adjustments, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(f"Employees: {len(employees)}")
    print(f"June records: {len(june_records)}")
    print(f"July records: {len(july_records)} (weekend defaults)")
    print(f"Position rates: {len(rates)}")
    print(f"Bonuses: {len(bonuses)}")
    print(f"Deductions: {len(deductions)}")
    print(f"Commission types: {len(commission_types)}")
    print(f"Payroll adjustments: {len(payroll_adjustments)}")
    print(f"Written to {OUT}")


if __name__ == "__main__":
    main()
