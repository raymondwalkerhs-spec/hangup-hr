"""Read net salaries from Excel Payroll_June2026 for reconciliation."""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "HR System June 2026 V.2 (1).xlsx"


def main():
    if not XLSX.exists():
        print("{}")
        return
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Payroll_June2026"]
    net_col = None
    for c in range(1, 70):
        if ws.cell(4, c).value == "Net Salary":
            net_col = c
            break
    out = {}
    for row in range(5, ws.max_row + 1):
        eid = ws.cell(row, 1).value
        if not eid:
            continue
        net = ws.cell(row, net_col).value if net_col else None
        try:
            out[str(eid).strip()] = float(net) if net is not None else None
        except (TypeError, ValueError):
            out[str(eid).strip()] = None
    print(json.dumps(out))
    wb.close()


if __name__ == "__main__":
    main()
