import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook(Path(__file__).parents[1] / "HR System June 2026 V.2 (1).xlsx", data_only=True)
ws = wb["Payroll_June2026"]
for c in range(1, 70):
    h = ws.cell(4, c).value
    if h:
        print(c, repr(h))
print("--- sample rows ---")
for row in range(5, 12):
    eid = ws.cell(row, 1).value
    extra = ws.cell(row, 14).value
    hold = ws.cell(row, 59).value
    net_col = None
    for c in range(1, 70):
        if ws.cell(4, c).value == "Net Salary":
            net_col = c
    net = ws.cell(row, net_col).value if net_col else None
    print(eid, "extra", extra, "hold", hold, "net", net)
if "Commission_Types" in wb.sheetnames:
    ct = wb["Commission_Types"]
    for r in range(1, min(15, ct.max_row + 1)):
        print("CT", [ct.cell(r, c).value for c in range(1, 6)])
wb.close()
