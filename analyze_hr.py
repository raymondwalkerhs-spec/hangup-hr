import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

path = r"H:\HR\HR System June 2026 V.2 (1).xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

formula_by_sheet = defaultdict(list)
sheet_refs = defaultdict(set)

for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                formula_by_sheet[name].append((cell.coordinate, formula))
                refs = re.findall(r"'([^']+)'!", formula)
                for r in refs:
                    sheet_refs[name].add(r)

lines = []
lines.append("=== FORMULA COUNT BY SHEET ===")
for name in wb.sheetnames:
    lines.append(f"{name}: {len(formula_by_sheet[name])} formulas")

lines.append("\n=== SHEET DEPENDENCIES ===")
for name in sorted(sheet_refs.keys()):
    refs = sorted(sheet_refs[name])
    if refs:
        lines.append(f"{name} -> {refs}")

lines.append("\n=== KEY FORMULA PATTERNS ===")
priority = [
    "Payroll_June2026", "Payroll_Template_May", "Attendance", "Employee_Database",
    "Bonus_Summary", "Bonus", "Deduction", "Cash", "Bank", "Insta", "Closed"
]
for name in wb.sheetnames:
    if name in priority or "Payroll" in name or "Bonus" in name or "Deduction" in name:
        formulas = formula_by_sheet[name]
        if formulas:
            lines.append(f"\n--- {name} ({len(formulas)} formulas) ---")
            patterns = defaultdict(int)
            for coord, f in formulas:
                pat = re.sub(r"\d+", "N", f)
                patterns[pat] += 1
            for pat, cnt in sorted(patterns.items(), key=lambda x: -x[1])[:15]:
                lines.append(f"  [{cnt}x] {pat[:350]}")

# Payroll full structure
ws = wb["Payroll_June2026"]
lines.append("\n=== PAYROLL_JUNE2026 ALL HEADERS (scanning rows 1-5) ===")
for row in range(1, 6):
    for col in range(1, min(80, ws.max_column + 1)):
        h = ws.cell(row=row, column=col).value
        if h:
            lines.append(f"R{row} {get_column_letter(col)}: {h}")

# Payroll formulas row 3-10
lines.append("\n=== PAYROLL SAMPLE FORMULAS ===")
for row in range(2, 12):
    for col in range(1, min(60, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            lines.append(f"{cell.coordinate}: {cell.value[:300]}")

# Attendance day columns
ws = wb["Attendance"]
lines.append("\n=== ATTENDANCE ROW 1-2 (date headers) ===")
for col in range(1, ws.max_column + 1):
    v1 = ws.cell(row=1, column=col).value
    v2 = ws.cell(row=2, column=col).value
    if v1 or v2:
        lines.append(f"{get_column_letter(col)}: R1={v1} | R2={v2}")

# FP Lateness formulas
lines.append("\n=== ATTENDANCE FP LATENESS FORMULAS (P3 sample) ===")
for col in range(16, min(60, ws.max_column + 1)):
    cell = ws.cell(row=3, column=col)
    if isinstance(cell.value, str) and cell.value.startswith("="):
        lines.append(f"{cell.coordinate}: {cell.value[:400]}")

# Bonus summary
if "Bonus_Summary" in wb.sheetnames:
    ws = wb["Bonus_Summary"]
    lines.append("\n=== BONUS_SUMMARY ===")
    for col in range(1, min(20, ws.max_column + 1)):
        lines.append(f"{get_column_letter(col)}: {ws.cell(row=1, column=col).value}")
    for row in range(2, 6):
        vals = [str(ws.cell(row=row, column=c).value) for c in range(1, min(12, ws.max_column + 1))]
        lines.append("  " + " | ".join(vals))

# Unit-specific bonus/deduction sheets
for sheet in ["Bonus_HS1", "Deduction_HS1", "Bonus_HS3", "Deduction_HS3"]:
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"\n=== {sheet} (first row headers) ===")
        for col in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=1, column=col).value
            if v:
                lines.append(f"  {get_column_letter(col)}: {v}")
        cnt = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value)
        lines.append(f"  Data rows: ~{cnt}")

# Employee database stats
ws = wb["Employee_Database"]
lines.append("\n=== EMPLOYEE DATABASE STATS ===")
statuses = defaultdict(int)
units = defaultdict(int)
positions = defaultdict(int)
for row in range(2, ws.max_row + 1):
    emp_id = ws.cell(row=row, column=1).value
    if not emp_id:
        continue
    status = ws.cell(row=row, column=7).value or "blank"
    unit = ws.cell(row=row, column=10).value or "blank"
    pos = ws.cell(row=row, column=8).value or "blank"
    statuses[status] += 1
    units[unit] += 1
    positions[pos] += 1
lines.append(f"Active employees with ID: {sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r, column=1).value)}")
lines.append("Statuses: " + str(dict(statuses)))
lines.append("Units: " + str(dict(units)))
lines.append("Top positions: " + str(dict(sorted(positions.items(), key=lambda x: -x[1])[:15])))

with open(r"H:\HR\formula_analysis.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("Done")
